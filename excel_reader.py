import pandas as pd
import openpyxl


df = pd.read_excel ('output/output.xlsx')
print (df)

# dataframe = openpyxl.load_workbook("output/output.xlsx")
# dataframe1 = dataframe.active

# for row in range(0, dataframe1.max_row):
#     for col in dataframe1.iter_cols(1, dataframe1.max_column):
#         print(col[row].value)

